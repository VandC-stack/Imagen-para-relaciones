#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de prueba para control_folios_anual.py
Demuestra diferentes casos de uso del generador de Excel
"""

import os
import sys
from datetime import datetime

def test_basic_generation():
    """Prueba básica: Generar reporte completo"""
    print("\n" + "="*70)
    print("TEST 1: Generación básica (todos los registros)")
    print("="*70)
    
    cmd = "python control_folios_anual.py -o /tmp/test_basico.xlsx"
    print(f"Comando: {cmd}")
    result = os.system(cmd)
    
    if result == 0:
        print("✅ Test 1 PASADO: Archivo generado exitosamente")
    else:
        print("❌ Test 1 FALLIDO: Error al generar archivo")
    
    return result == 0

def test_date_filtering():
    """Prueba con filtrado de fechas"""
    print("\n" + "="*70)
    print("TEST 2: Filtrado por rango de fechas (Noviembre 2025)")
    print("="*70)
    
    cmd = "python control_folios_anual.py -o /tmp/test_fechas.xlsx -fi 2025-11-01 -ff 2025-11-30"
    print(f"Comando: {cmd}")
    result = os.system(cmd)
    
    if result == 0:
        print("✅ Test 2 PASADO: Filtro de fechas aplicado correctamente")
    else:
        print("❌ Test 2 FALLIDO: Error al aplicar filtro de fechas")
    
    return result == 0

def test_custom_output():
    """Prueba con nombre personalizado"""
    print("\n" + "="*70)
    print("TEST 3: Nombre de archivo personalizado")
    print("="*70)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"/tmp/Control_Folios_{timestamp}.xlsx"
    cmd = f"python control_folios_anual.py -o {output_file}"
    print(f"Comando: {cmd}")
    result = os.system(cmd)
    
    if result == 0 and os.path.exists(output_file):
        size = os.path.getsize(output_file)
        print(f"✅ Test 3 PASADO: Archivo creado: {output_file}")
        print(f"   Tamaño: {size} bytes")
    else:
        print("❌ Test 3 FALLIDO: Archivo no creado")
    
    return result == 0 and os.path.exists(output_file)

def test_help():
    """Prueba del comando de ayuda"""
    print("\n" + "="*70)
    print("TEST 4: Comando de ayuda")
    print("="*70)
    
    cmd = "python control_folios_anual.py --help"
    print(f"Comando: {cmd}")
    result = os.system(cmd)
    
    if result == 0:
        print("✅ Test 4 PASADO: Ayuda mostrada correctamente")
    else:
        print("❌ Test 4 FALLIDO: Error al mostrar ayuda")
    
    return result == 0

def verify_excel_structure():
    """Verifica la estructura del Excel generado"""
    print("\n" + "="*70)
    print("TEST 5: Verificación de estructura del Excel")
    print("="*70)
    
    try:
        import openpyxl
        
        # Generar archivo de prueba
        test_file = "/tmp/test_estructura.xlsx"
        os.system(f"python control_folios_anual.py -o {test_file} > /dev/null 2>&1")
        
        # Cargar y verificar
        wb = openpyxl.load_workbook(test_file)
        ws = wb.active
        
        # Verificar encabezados esperados
        expected_headers = [
            "NÚMERO DE SOLICITUD",
            "CLIENTE",
            "NÚMERO DE CONTRATO",
            "RFC",
            "CURP",
            "PRODUCTO VERIFICADO",
            "MARCAS",
            "NOM",
            "TIPO DE DOCUMENTO OFICIAL EMITIDO",
            "DOCUMENTO EMITIDO",
            "FECHA DE DOCUMENTO EMITIDO",
            "VERIFICADOR",
            "PEDIMENTO DE IMPORTACION",
            "FECHA DE DESADUANAMIENTO (CUANDO APLIQUE)",
            "FECHA DE VISITA (CUANDO APLIQUE)",
            "MODELOS",
            "SOL EMA",
            "FOLIO EMA",
            "INSP EMA"
        ]
        
        print(f"📊 Estadísticas del archivo:")
        print(f"   - Filas: {ws.max_row}")
        print(f"   - Columnas: {ws.max_column}")
        
        # Verificar encabezados
        all_headers_match = True
        for i, expected in enumerate(expected_headers, 1):
            actual = ws.cell(1, i).value
            if actual != expected:
                print(f"❌ Encabezado columna {i}: esperado '{expected}', obtenido '{actual}'")
                all_headers_match = False
        
        if all_headers_match:
            print(f"✅ Todos los encabezados coinciden ({len(expected_headers)} columnas)")
        
        # Verificar que hay datos
        has_data = ws.max_row > 1
        if has_data:
            print(f"✅ El archivo contiene datos ({ws.max_row - 1} registros)")
        else:
            print("❌ El archivo no contiene datos")
        
        # Verificar estilos de encabezados
        first_header = ws.cell(1, 1)
        has_style = first_header.font.bold and first_header.fill.start_color.rgb
        if has_style:
            print("✅ Los encabezados tienen formato aplicado")
        else:
            print("❌ Los encabezados no tienen formato")
        
        success = all_headers_match and has_data and has_style
        
        if success:
            print("\n✅ Test 5 PASADO: Estructura del Excel correcta")
        else:
            print("\n❌ Test 5 FALLIDO: Problemas con la estructura del Excel")
        
        return success
        
    except ImportError:
        print("❌ Test 5 OMITIDO: openpyxl no está instalado")
        return False
    except Exception as e:
        print(f"❌ Test 5 FALLIDO: {e}")
        return False

def main():
    """Ejecutar todos los tests"""
    print("\n" + "="*70)
    print("🧪 SUITE DE PRUEBAS - control_folios_anual.py")
    print("="*70)
    
    results = []
    
    # Ejecutar tests
    results.append(("Generación básica", test_basic_generation()))
    results.append(("Filtrado por fechas", test_date_filtering()))
    results.append(("Nombre personalizado", test_custom_output()))
    results.append(("Comando de ayuda", test_help()))
    results.append(("Estructura del Excel", verify_excel_structure()))
    
    # Resumen
    print("\n" + "="*70)
    print("📊 RESUMEN DE PRUEBAS")
    print("="*70)
    
    total = len(results)
    passed = sum(1 for _, result in results if result)
    
    for name, result in results:
        status = "✅ PASADO" if result else "❌ FALLIDO"
        print(f"{status}: {name}")
    
    print("\n" + "="*70)
    print(f"Total: {passed}/{total} pruebas pasadas")
    
    if passed == total:
        print("🎉 ¡Todos los tests pasaron exitosamente!")
        return 0
    else:
        print(f"⚠️  {total - passed} test(s) fallaron")
        return 1

if __name__ == "__main__":
    exit(main())
