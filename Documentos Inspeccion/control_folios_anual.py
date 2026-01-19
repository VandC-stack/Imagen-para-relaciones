import json
import os
import re
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import unicodedata
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ControlFoliosAnual:
    """Clase para generar el control de folios anual desde archivos JSON"""
    
    def __init__(self, data_dir: str = "data"):
        """
        Inicializar el generador de control de folios
        
        Args:
            data_dir: Directorio donde se encuentran los archivos JSON
        """
        self.data_dir = data_dir
        self.clientes = []
        self.firmas = []
        self.tabla_relacion = []
        self.historial_visitas = []
        self.folio_to_cliente = {}  # Mapeo de folio a cliente
        self.normas = []
        self._dictamen_cache = {}
        # Mapeo (SOLICITUD, FOLIO) -> FECHA DE ENTRADA extraída de backups
        self.backup_fecha_entrada: Dict[Tuple[str, str], str] = {}
        
    def cargar_datos(self) -> Tuple[bool, str]:
        """
        Cargar todos los archivos JSON necesarios
        
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            # Cargar Clientes.json
            clientes_path = os.path.join(self.data_dir, "Clientes.json")
            if os.path.exists(clientes_path):
                with open(clientes_path, 'r', encoding='utf-8') as f:
                    self.clientes = json.load(f)
                print(f"✅ Clientes cargados: {len(self.clientes)} registros")
            else:
                print(f"⚠️ Advertencia: No se encontró {clientes_path}. Continuando con clientes vacíos.")
                self.clientes = []
            
            # Cargar Firmas.json
            firmas_path = os.path.join(self.data_dir, "Firmas.json")
            if os.path.exists(firmas_path):
                with open(firmas_path, 'r', encoding='utf-8') as f:
                    self.firmas = json.load(f)
                print(f"✅ Firmas cargadas: {len(self.firmas)} registros")
            else:
                print(f"⚠️ Advertencia: No se encontró {firmas_path}. Continuando con firmas vacías.")
                self.firmas = []
            
            # Cargar tabla_de_relacion.json
            tabla_path = os.path.join(self.data_dir, "tabla_de_relacion.json")
            if os.path.exists(tabla_path):
                with open(tabla_path, 'r', encoding='utf-8') as f:
                    self.tabla_relacion = json.load(f)
                print(f"✅ Tabla de relación cargada: {len(self.tabla_relacion)} registros")
            else:
                print(f"⚠️ Advertencia: No se encontró {tabla_path}. Continuando con tabla_de_relacion vacía.")
                self.tabla_relacion = []

            # Cargar backups de tabla_relacion (cada visita genera una copia aquí)
            # y construir un mapeo (SOLICITUD, FOLIO) -> FECHA DE ENTRADA
            try:
                backups_dir = os.path.join(self.data_dir, "tabla_relacion_backups")
                if os.path.exists(backups_dir):
                    for bf in os.listdir(backups_dir):
                        if not bf.lower().endswith('.json'):
                            continue
                        bfp = os.path.join(backups_dir, bf)
                        try:
                            with open(bfp, 'r', encoding='utf-8') as bf_f:
                                bdata = json.load(bf_f)
                        except Exception:
                            continue

                        # normalizar lista de registros si es dict o lista
                        records = []
                        if isinstance(bdata, dict):
                            # buscar primer value que sea lista
                            for v in bdata.values():
                                if isinstance(v, list):
                                    records = v
                                    break
                            if not records:
                                # fallback: si el dict parece ya una lista por claves numéricas
                                try:
                                    # intentar interpretar como lista de diccionarios
                                    possible = list(bdata.values())
                                    if all(isinstance(x, dict) for x in possible):
                                        records = possible
                                except Exception:
                                    records = []
                        elif isinstance(bdata, list):
                            records = bdata

                        for r in records:
                            sol = str(r.get('SOLICITUD', '')).strip()
                            fol = str(r.get('FOLIO', '')).strip()
                            fecha_ent = r.get('FECHA DE ENTRADA') or r.get('FECHA_DE_ENTRADA')
                            if sol and fol and fecha_ent:
                                key = (sol, fol)
                                # preferir la primera encontrada (no sobrescribir)
                                if key not in self.backup_fecha_entrada:
                                    self.backup_fecha_entrada[key] = str(fecha_ent)
                else:
                    # si no existe carpeta de backups, seguir silenciosamente
                    pass
            except Exception:
                # no bloquear la carga por errores en backups
                pass

            # Cargar Normas.json (opcional, para mostrar nombres completos de NOM)
            normas_path = os.path.join(self.data_dir, "Normas.json")
            if os.path.exists(normas_path):
                with open(normas_path, 'r', encoding='utf-8') as f:
                    try:
                        self.normas = json.load(f)
                        print(f"✅ Normas cargadas: {len(self.normas)}")
                    except Exception:
                        self.normas = []
            
            # Cargar historial_visitas.json (opcional, para mapeo de clientes)
            historial_path = os.path.join(self.data_dir, "historial_visitas.json")
            if os.path.exists(historial_path):
                with open(historial_path, 'r', encoding='utf-8') as f:
                    hist_data = json.load(f)
                    if isinstance(hist_data, dict) and 'visitas' in hist_data:
                        self.historial_visitas = hist_data['visitas']
                        # Crear mapeo de folio a cliente
                        self._crear_mapeo_folio_cliente()
                        print(f"✅ Historial de visitas cargado: {len(self.historial_visitas)} registros")
            
            return True, "Datos cargados correctamente"
            
        except json.JSONDecodeError as e:
            return False, f"Error al decodificar JSON: {e}"
        except Exception as e:
            return False, f"Error al cargar datos: {e}"
    
    def _crear_mapeo_folio_cliente(self):
        """
        Crear un mapeo entre folios y clientes desde el historial de visitas
        """
        for visita in self.historial_visitas:
            cliente_nombre = visita.get('cliente', '')
            folios_str = visita.get('folios_utilizados', '')
            
            if not cliente_nombre or not folios_str:
                continue
            
            # Parse folio range (e.g., "075339 - 075552")
            if ' - ' in folios_str:
                parts = folios_str.split(' - ')
                if len(parts) == 2:
                    try:
                        inicio = int(parts[0].strip())
                        fin = int(parts[1].strip())

                        # Map all folios in range to this client
                        for folio_num in range(inicio, fin + 1):
                            self.folio_to_cliente[folio_num] = cliente_nombre
                    except ValueError:
                        pass
    
    def buscar_cliente_por_solicitud(self, solicitud: str, folio: int) -> Optional[Dict]:
        """
        Buscar información del cliente basándose en el folio
        
        Args:
            solicitud: Número de solicitud (ej: "006916/25") - no usado directamente
            folio: Número de folio para buscar el cliente
            
        Returns:
            Diccionario con información del cliente o None
        """
        # Primero intentar buscar por folio en el historial
        cliente_nombre = self.folio_to_cliente.get(folio)
        
        if cliente_nombre:
            # Buscar información completa del cliente por nombre
            for cliente in self.clientes:
                if cliente.get('CLIENTE', '').strip().upper() == cliente_nombre.strip().upper():
                    return cliente
        
        # Si no se encontró, retornar información genérica con el nombre del historial
        if cliente_nombre:
            return {
                'CLIENTE': cliente_nombre,
                'NÚMERO_DE_CONTRATO': 'N/A',
                'RFC': 'N/A',
                'CURP': 'N/A'
            }
        
        # Como último recurso, retornar N/A
        return {
            'CLIENTE': 'N/A',
            'NÚMERO_DE_CONTRATO': 'N/A',
            'RFC': 'N/A',
            'CURP': 'N/A'
        }
    
    def buscar_inspector_por_firma(self, firma: str) -> str:
        """
        Buscar el nombre completo del inspector por su firma
        
        Args:
            firma: Código de firma (ej: "GRAMIREZ")
            
        Returns:
            Nombre completo del inspector o "N/A"
        """
        for inspector in self.firmas:
            if inspector.get("FIRMA") == firma:
                # Intentar extraer nombre por varias claves posibles
                nombre = (
                    inspector.get("NOMBRE") or inspector.get("NOMBRE_COMPLETO")
                    or inspector.get("NOMBRE DE INSPECTOR") or inspector.get("nombre")
                    or inspector.get("NOMBRE INSPECTOR") or inspector.get("NOMBRE_COMPLETO_INSPECTOR")
                    or inspector.get("FIRMA")
                )
                return self._normalize_name(nombre)
        return "N/A"

    def _normalize_name(self, name: Optional[str]) -> str:
        """
        Normalizar nombre: convertir a mayúsculas y quitar acentos.
        """
        if not name:
            return "N/A"
        nfkd = unicodedata.normalize('NFKD', str(name))
        only_ascii = ''.join([c for c in nfkd if not unicodedata.combining(c)])
        return only_ascii.upper()
    
    def formatear_folio_ema(self, folio) -> str:
        """
        Formatear el folio EMA a 6 dígitos
        
        Args:
            folio: Número de folio
            
        Returns:
            Folio formateado a 6 dígitos
        """
        try:
            folio_str = str(int(folio))
            return folio_str.zfill(6)
        except (ValueError, TypeError):
            return "000000"
    
    def extraer_sol_ema(self, numero_solicitud: str) -> str:
        """
        Extraer los últimos valores del número de solicitud
        
        Args:
            numero_solicitud: Número de solicitud completo (ej: "006916/25")
            
        Returns:
            Últimos valores separados por guión
        """
        if not numero_solicitud:
            return "N/A"
        # Extraer los últimos componentes separados por '/'
        partes = numero_solicitud.split('/')
        if len(partes) >= 2:
            # construir sol-xxx, pero quitar sufijo de año si el último componente es año corto (ej '25')
            last = partes[-1]
            penult = partes[-2]
            if last.isdigit() and len(last) == 2:
                return penult
            return f"{penult}-{last}"
        return numero_solicitud

    def _find_dictamen(self, solicitud: str, folio) -> Optional[Dict]:
        """Buscar dictamen JSON en data/Dictamenes que coincida con solicitud y folio."""
        try:
            # Normalizar folio y solicitud para mejorar coincidencias
            folio_s = str(folio).strip() if folio is not None else ''
            sol_search = str(solicitud).strip() if solicitud is not None else ''
            # Si viene con formato 'XXXX/25', usar la parte antes de la barra
            sol_base = sol_search.split('/')[0] if '/' in sol_search else sol_search

            dicts_dir = os.path.join(self.data_dir, 'Dictamenes')
            if dicts_dir in self._dictamen_cache:
                files = self._dictamen_cache[dicts_dir]
            else:
                if not os.path.exists(dicts_dir):
                    return None
                # Index both JSON and PDF files so we can match dictamen info even
                # when only a PDF exists for a folio.
                files = [os.path.join(dicts_dir, f) for f in os.listdir(dicts_dir)
                         if f.lower().endswith('.json') or f.lower().endswith('.pdf')]
                self._dictamen_cache[dicts_dir] = files

            # Buscamos coincidencias estrictas: preferir igualdad de folio+solicitud.
            for fp in files:
                fname = os.path.basename(fp)
                # If JSON, load; if PDF, build a minimal dictamen object from filename
                d = None
                if fname.lower().endswith('.json'):
                    try:
                        with open(fp, 'r', encoding='utf-8') as f:
                            d = json.load(f)
                    except Exception:
                        continue
                elif fname.lower().endswith('.pdf'):
                    # create minimal dict structure from filename
                    name_base = os.path.splitext(fname)[0]
                    digit_sequences = re.findall(r"(\d+)", name_base)
                    fol_guess = ''
                    sol_guess = ''
                    if digit_sequences:
                        fol_guess = max(digit_sequences, key=lambda s: len(s))
                    sol_match = re.search(r"([A-Za-z0-9\-]{4,})", name_base)
                    if sol_match:
                        sol_guess = sol_match.group(1)
                    d = {
                        'identificacion': {
                            'solicitud': sol_guess,
                            'folio': fol_guess,
                            'cadena_identificacion': ''
                        }
                    }
                else:
                    continue

                ident = d.get('identificacion', {})
                sol_file = str(ident.get('solicitud', '')).strip()
                fol_file = str(ident.get('folio', '')).strip()
                cadena = (ident.get('cadena_identificacion') or '')

                # 1) Coincidencia exacta en folio y solicitud (cuando ambos estén presentes)
                if fol_file and folio_s:
                    equal_folio = False
                    try:
                        # Comparar como texto (posible padded) y como dígitos (ignorar ceros)
                        if fol_file == folio_s:
                            equal_folio = True
                        else:
                            digits_a = ''.join(ch for ch in fol_file if ch.isdigit())
                            digits_b = ''.join(ch for ch in folio_s if ch.isdigit())
                            if digits_a and digits_b and int(digits_a) == int(digits_b):
                                equal_folio = True
                    except Exception:
                        equal_folio = False
                    if equal_folio:
                        if sol_base:
                            if sol_file and (sol_file == sol_base or sol_file.endswith(sol_base)):
                                return d
                            # También checar si la solicitud completa aparece en la cadena_identificacion
                            if cadena and sol_base in cadena:
                                return d
                            # también intentar por filename (considerar variantes sin padding)
                            fname_digits = ''.join(ch for ch in fname if ch.isdigit())
                            if (f"_{folio_s}_" in fname and sol_base and f"_{sol_base}_" in fname) or (digits_a and str(int(digits_a)) in fname_digits):
                                return d
                        else:
                            return d

                # 2) Coincidencia exacta en solicitud cuando folio no disponible
                if sol_base and sol_file and (sol_file == sol_base or sol_file.endswith(sol_base)):
                    return d

                # 3) Intentar por filename patrones (folio y solicitud dentro del nombre)
                try:
                    if folio_s and f"_{folio_s}_" in fname:
                        if sol_base:
                            if f"_{sol_base}_" in fname:
                                return d
                            # si filename contiene folio pero no solicitud, aún puede ser válido
                            return d
                except Exception:
                    pass

                # 4) Por último, coincidencia en cadena_identificacion pero más restrictiva: buscar sol_base rodeada
                if cadena and sol_base:
                    if sol_base in cadena:
                        return d
            # Si no encontramos ninguna coincidencia estricta, devolvemos None
        except Exception:
            return None
        return None

    def _lookup_backup_fecha(self, solicitud: str, folio) -> Optional[str]:
        """
        Buscar una FECHA DE ENTRADA en el mapeo de backups usando varias heurísticas.
        """
        try:
            sol = str(solicitud).strip() if solicitud is not None else ''
            fol_s = str(folio).strip() if folio is not None else ''

            # intento 1: búsqueda exacta
            key = (sol, fol_s)
            if key in self.backup_fecha_entrada:
                return self.backup_fecha_entrada[key]

            # intento 2: buscar con la parte antes de '/' en la solicitud
            sol_base = sol.split('/')[0] if '/' in sol else sol
            for (k_sol, k_fol), fecha in self.backup_fecha_entrada.items():
                if k_sol == sol_base and (not fol_s or k_fol == fol_s):
                    return fecha

            # intento 3: buscar por folio solamente
            if fol_s:
                for (k_sol, k_fol), fecha in self.backup_fecha_entrada.items():
                    if k_fol == fol_s:
                        return fecha

            # intento 4: buscar por coincidencia parcial en solicitud
            for (k_sol, k_fol), fecha in self.backup_fecha_entrada.items():
                if sol and sol in k_sol:
                    return fecha

        except Exception:
            return None
        return None
    
    def agrupar_por_dictamen(self) -> List[Dict]:
        """
        Agrupar los registros de tabla_relacion por dictamen (SOLICITUD + FOLIO)
        
        Returns:
            Lista de dictámenes agrupados con su información
        """
        dictamenes = {}
        
        for registro in self.tabla_relacion:
            solicitud = registro.get("SOLICITUD", "")
            folio = registro.get("FOLIO", "")
            
            # Crear clave única por dictamen
            clave_dictamen = f"{solicitud}_{folio}"
            
            if clave_dictamen not in dictamenes:
                dictamenes[clave_dictamen] = {
                    "solicitud": solicitud,
                    "folio": folio,
                    "registros": []
                }
            
            dictamenes[clave_dictamen]["registros"].append(registro)
        
        return list(dictamenes.values())
    
    def generar_fila_excel(self, dictamen: Dict) -> Dict:
        """
        Generar una fila del Excel a partir de un dictamen
        
        Args:
            dictamen: Diccionario con información del dictamen
            
        Returns:
            Diccionario con los datos de la fila
        """
        registros = dictamen["registros"]
        primer_registro = registros[0] if registros else {}
        
        solicitud = dictamen["solicitud"]
        folio = dictamen["folio"]
        
        # Buscar cliente usando el folio
        try:
            folio_num = int(folio) if folio else 0
        except (ValueError, TypeError):
            folio_num = 0
        
        cliente = self.buscar_cliente_por_solicitud(solicitud, folio_num)

        # Intentar localizar dictamen JSON para extraer cadena_identificacion y norma completa
        cadena_ident = None
        norma_codigo = None
        # Si el dictamen ya lleva el JSON embebido (creado en crear_excel), úsalo
        if dictamen.get('dictamen_json'):
            dictamen_json = dictamen.get('dictamen_json')
            ident = dictamen_json.get('identificacion', {})
            cadena_ident = ident.get('cadena_identificacion')
            norma = dictamen_json.get('norma', {})
            norma_codigo = norma.get('codigo') or norma.get('NOM')
        else:
            dictamen_json = self._find_dictamen(solicitud, folio)
            if dictamen_json:
                ident = dictamen_json.get('identificacion', {})
                cadena_ident = ident.get('cadena_identificacion')
                norma = dictamen_json.get('norma', {})
                norma_codigo = norma.get('codigo') or norma.get('NOM')
        
        # Obtener información del inspector
        firma = primer_registro.get("FIRMA", "")
        nombre_inspector = self.buscar_inspector_por_firma(firma)
        
        # Extraer descripciones, marcas, NOMs y modelos de todos los registros
        descripciones = set()
        marcas = set()
        noms = set()

        # Preferir los códigos que vienen en el JSON del dictamen (tabla_productos)
        modelos = []
        if dictamen.get('dictamen_json'):
            try:
                tp = dictamen['dictamen_json'].get('tabla_productos', [])
                for p in tp:
                    c = p.get('codigo')
                    if c is not None:
                        modelos.append(str(c))
                # También extraer marcas y descripciones desde el JSON si no vienen en registros
                prod = dictamen['dictamen_json'].get('producto', {})
                if prod.get('descripcion'):
                    descripciones.add(prod.get('descripcion'))
                if tp and isinstance(tp, list):
                    for p in tp:
                        m = p.get('marca')
                        if m:
                            marcas.add(m)
            except Exception:
                modelos = []

        # Si no hay dictamen_json, o además, recorrer registros (tabla_relacion) para completar datos
        for reg in registros:
            if reg.get("DESCRIPCION"):
                descripciones.add(reg.get("DESCRIPCION"))
            if reg.get("MARCA"):
                marcas.add(reg.get("MARCA"))
            if reg.get("CLASIF UVA"):
                noms.add(str(reg.get("CLASIF UVA")))

            # Aceptar tanto 'CODIGO' simple como 'CODIGOS' lista en registros de tabla_relacion
            if reg.get("CODIGO"):
                val = reg.get("CODIGO")
                if isinstance(val, (list, tuple)):
                    for v in val:
                        modelos.append(str(v))
                else:
                    if isinstance(val, str) and ',' in val:
                        for v in [x.strip() for x in val.split(',') if x.strip()]:
                            modelos.append(str(v))
                    else:
                        modelos.append(str(val))
            if reg.get("CODIGOS"):
                try:
                    for c in reg.get("CODIGOS"):
                        modelos.append(str(c))
                except Exception:
                    pass
        
        # Preparar valores derivados
        numero_solicitud_display = None
        if cadena_ident:
            # Intentar extraer token después de 'Solicitud de Servicio:'
            m = re.search(r"Solicitud de Servicio:\s*([A-Za-z0-9\-]+)", cadena_ident)
            if m:
                numero_solicitud_display = m.group(1)
            else:
                m2 = re.search(r"([A-Za-z0-9\-]+-[0-9]+)$", cadena_ident)
                if m2:
                    numero_solicitud_display = m2.group(1)
                else:
                    numero_solicitud_display = cadena_ident

        numero_solicitud_display = numero_solicitud_display or solicitud or "N/A"

        # Tipo de documento (mapear letra a texto)
        tipo_raw = primer_registro.get("TIPO DE DOCUMENTO") or primer_registro.get("TIPO DE DOCUMENTO OFICIAL EMITIDO", "D")
        tipo_display = "Dictamen" if str(tipo_raw).strip().upper() == 'D' else str(tipo_raw)

        # NOM: preferir norma del dictamen, si no mapear CLASIF UVA usando Normas.json
        if norma_codigo:
            nom_display = norma_codigo
        else:
            mapped = []
            for c in noms:
                mapped_nom = None
                try:
                    ci = int(c)
                    padded = f"{ci:03d}"
                except Exception:
                    padded = str(c)
                for n in self.normas:
                    nom_field = n.get('NOM', '')
                    if padded and padded in nom_field:
                        mapped_nom = nom_field
                        break
                    if str(c) and str(c) in nom_field:
                        mapped_nom = nom_field
                        break
                if mapped_nom:
                    mapped.append(mapped_nom)
                else:
                    mapped.append(str(c))
            nom_display = ", ".join(sorted(set(mapped))) if mapped else "N/A"

        # Para la columna DOCUMENTO EMITIDO preferimos usar un token informativo
        # extraído de `cadena_identificacion`. Normalmente la cadena contiene
        # un bloque "Solicitud de Servicio: <token>" que es el más completo. Si
        # no, tomamos el primer token (antes de '/') y, cuando disponemos de
        # `solicitud` y `folio`, anexamos la parte '<solicitud>-<folio>' para
        # obtener formatos como '25049USDNOM-141-SSA1007045-126'.
        if cadena_ident:
            try:
                token = None
                # Preferir token después de 'Solicitud de Servicio:' si existe
                msol = re.search(r"Solicitud de Servicio:\s*([A-Za-z0-9\-/]+)", str(cadena_ident))
                if msol:
                    token = msol.group(1).split('/')[0]
                else:
                    # Fallback: primer token antes de espacios, y quitar cualquier parte tras '/'
                    token = str(cadena_ident).strip().split()[0].split('/')[0]

                sol_part = self.extraer_sol_ema(solicitud) or ''
                # Build a sensible documento_emitido without duplicating fragments.
                if sol_part and folio:
                    suffix = f"{sol_part}-{folio}".lstrip('-')
                    # If token already contains the suffix or the sol_part, avoid appending again
                    if token and (token.endswith(suffix) or sol_part in token):
                        documento_emitido = token
                    else:
                        # ensure separator
                        sep = ''
                        if token and not token.endswith('-'):
                            sep = '-'
                        documento_emitido = f"{token}{sep}{suffix}" if token else f"{suffix}"
                else:
                    documento_emitido = token
            except Exception:
                documento_emitido = numero_solicitud_display
        else:
            documento_emitido = numero_solicitud_display

        # Normalizar modelos: mantener orden y eliminar duplicados vacíos
        modelos_norm = []
        for m in modelos:
            if not m:
                continue
            if m not in modelos_norm:
                modelos_norm.append(m)

        # Preferir FECHA DE ENTRADA desde backups (tabla_relacion_backups) si existe
        fecha_desaduanamiento = None
        try:
            fecha_desaduanamiento = self._lookup_backup_fecha(solicitud, folio)
            if not fecha_desaduanamiento:
                # buscar en registros cualquiera con FECHA DE ENTRADA válida
                for r in registros:
                    fe = r.get('FECHA DE ENTRADA') or r.get('FECHA_DE_ENTRADA')
                    if fe:
                        fecha_desaduanamiento = fe
                        break
        except Exception:
            fecha_desaduanamiento = None

        fila = {
            "NÚMERO DE SOLICITUD": numero_solicitud_display,
            "CLIENTE": cliente.get("CLIENTE", "N/A") if cliente else "N/A",
            "NÚMERO DE CONTRATO": cliente.get("NÚMERO_DE_CONTRATO", "N/A") if cliente else "N/A",
            "RFC": cliente.get("RFC", "N/A") if cliente else "N/A",
            "CURP": "N/A",
            "PRODUCTO VERIFICADO": ", ".join(descripciones) if descripciones else "N/A",
            "MARCAS": ", ".join(marcas) if marcas else "N/A",
            "NOM": nom_display,
            "TIPO DE DOCUMENTO OFICIAL EMITIDO": tipo_display,
            "DOCUMENTO EMITIDO": documento_emitido or "N/A",
            "FECHA DE DOCUMENTO EMITIDO": primer_registro.get("FECHA DE EMISION DE SOLICITUD", "N/A"),
            "VERIFICADOR": self._normalize_name(nombre_inspector),
            "PEDIMENTO DE IMPORTACION": primer_registro.get("PEDIMENTO", "N/A"),
            "FECHA DE DESADUANAMIENTO (CUANDO APLIQUE)": fecha_desaduanamiento or primer_registro.get("FECHA DE ENTRADA", "N/A"),
            "FECHA DE VISITA (CUANDO APLIQUE)": primer_registro.get("FECHA DE VERIFICACION", "N/A"),
            "MODELOS": ", ".join(modelos_norm) if modelos_norm else "N/A",
            "SOL EMA": self.extraer_sol_ema(solicitud),
            "FOLIO EMA": self.formatear_folio_ema(folio),
            "INSP EMA": self._normalize_name(nombre_inspector)
        }
        
        return fila
    
    def filtrar_por_fechas(self, fila: Dict, fecha_inicio: Optional[str] = None, 
                          fecha_fin: Optional[str] = None) -> bool:
        """
        Filtrar una fila por rango de fechas
        
        Args:
            fila: Fila de datos
            fecha_inicio: Fecha de inicio en formato YYYY-MM-DD
            fecha_fin: Fecha de fin en formato YYYY-MM-DD
            
        Returns:
            True si la fila está en el rango, False si no
        """
        if not fecha_inicio and not fecha_fin:
            return True
        
        # Usar la fecha de verificación para filtrar. Si no existe, intentar otros campos.
        fecha_str = fila.get("FECHA DE VISITA (CUANDO APLIQUE)", "")

        # Si no hay fecha de visita, intentar usar fecha de documento emitido o fecha de desaduanamiento
        if not fecha_str or fecha_str == "N/A":
            fecha_str = fila.get("FECHA DE DOCUMENTO EMITIDO") or fila.get("FECHA DE DESADUANAMIENTO (CUANDO APLIQUE)") or fila.get("FECHA DE EMISION DE SOLICITUD")

        # Si aún no hay fecha, incluir el registro (comportamiento más permisivo)
        if not fecha_str or fecha_str == "N/A":
            return True
        
        try:
            # Intentar parsear la fecha en diferentes formatos
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"]:
                try:
                    fecha = datetime.strptime(fecha_str, fmt)
                    break
                except ValueError:
                    continue
            else:
                # Si no se pudo parsear, incluir el registro
                return True
            
            if fecha_inicio:
                inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
                if fecha < inicio:
                    return False
            
            if fecha_fin:
                fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
                if fecha > fin:
                    return False
            
            return True
            
        except Exception:
            # En caso de error, incluir el registro
            return True
    
    def crear_excel(self, nombre_archivo: str, fecha_inicio: Optional[str] = None,
                   fecha_fin: Optional[str] = None) -> Tuple[bool, str]:
        """
        Crear el archivo Excel con el control de folios
        
        Args:
            nombre_archivo: Nombre del archivo Excel a crear
            fecha_inicio: Fecha de inicio para filtrar (YYYY-MM-DD)
            fecha_fin: Fecha de fin para filtrar (YYYY-MM-DD)
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            print("\n🚀 Generando archivo Excel...")
            
            # Crear libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Control de Folios"
            
            # Definir encabezados
            encabezados = [
                "NÚMERO DE SOLICITUD",
                "CLIENTE",
                "NÚMERO DE CONTRATO",
                "RFC",
                "CURP",
                "PRODUCTO VERIFICADO",
                "MARCAS",
                "NOM",
                "TIPO DE DOCUMENTO OFICIAL EMITIDO",
                "DOCUMENTO EMITIDO",
                "FECHA DE DOCUMENTO EMITIDO",
                "VERIFICADOR",
                "PEDIMENTO DE IMPORTACION",
                "FECHA DE DESADUANAMIENTO (CUANDO APLIQUE)",
                "FECHA DE VISITA (CUANDO APLIQUE)",
                "MODELOS",
                "SOL EMA",
                "FOLIO EMA",
                "INSP EMA"
            ]
            
            # Escribir encabezados
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col, value=encabezado)
                # Estilo para encabezados
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                celda.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Construir dictamenes a partir de los archivos JSON en data/Dictamenes
            # Usamos como clave única el token de 'Solicitud de Servicio' extraído
            # desde 'cadena_identificacion' (ej: 25049USDNOM-004-SE-2021006916-1).
            dictamenes_map: Dict[str, Dict] = {}
            try:
                dicts_dir = os.path.join(self.data_dir, 'Dictamenes')
                if os.path.exists(dicts_dir):
                    for fname in os.listdir(dicts_dir):
                        fp = os.path.join(dicts_dir, fname)
                        # Si es JSON, cargar como antes
                        if fname.lower().endswith('.json'):
                            try:
                                with open(fp, 'r', encoding='utf-8') as f:
                                    d = json.load(f)
                            except Exception:
                                continue

                            ident = d.get('identificacion', {})
                            sol = str(ident.get('solicitud', '')).strip()
                            fol = str(ident.get('folio', '')).strip()
                            cadena = str(ident.get('cadena_identificacion') or '').strip()
                        # Si es PDF u otro binario con extensión .pdf, crear registro mínimo a partir del nombre
                        elif fname.lower().endswith('.pdf'):
                            # Extraer tokens del nombre de archivo
                            name_base = os.path.splitext(fname)[0]
                            cadena = ''
                            sol = ''
                            fol = ''

                            # Buscar secuencias de dígitos (tomar la más larga como folio probable)
                            digit_sequences = re.findall(r"(\d+)", name_base)
                            if digit_sequences:
                                # elegir la secuencia más larga (probablemente el folio)
                                fol_candidate = max(digit_sequences, key=lambda s: len(s))
                                fol = fol_candidate

                            # Intentar extraer una solicitud/token alfanumérico (patrón similar a 'XXXX-123' o 'XXXXX')
                            sol_match = re.search(r"([A-Za-z0-9\-]{4,})", name_base)
                            if sol_match:
                                sol = sol_match.group(1)

                            # Construir un objeto mínimo que siga la estructura esperada
                            d = {
                                'identificacion': {
                                    'solicitud': sol,
                                    'folio': fol,
                                    'cadena_identificacion': cadena
                                },
                                'fechas': {},
                                'producto': {},
                                'tabla_productos': [],
                                'firmas': {}
                            }
                        else:
                            continue

                        # Extraer token 'Solicitud de Servicio' preferido
                        sol_token = None
                        if cadena:
                            m = re.search(r"Solicitud de Servicio:\s*([A-Za-z0-9\-]+)", cadena)
                            if m:
                                sol_token = m.group(1)
                            else:
                                m2 = re.search(r"([A-Za-z0-9\-]+-[0-9]+)$", cadena)
                                if m2:
                                    sol_token = m2.group(1)

                        # Clave principal: usar siempre 'solicitud_folio' para consistencia
                        key = f"{sol}_{fol}"
                        # Guardar también sol_token si se extrajo (como dato auxiliar)

                        # Crear registro mínimo basado en el JSON completo
                        registro = {}
                        fechas = d.get('fechas', {})
                        producto = d.get('producto', {})
                        tabla_prod = d.get('tabla_productos', [])
                        firmas = d.get('firmas', {})

                        registro['FECHA DE EMISION DE SOLICITUD'] = fechas.get('emision')
                        registro['FECHA DE VERIFICACION'] = fechas.get('verificacion')
                        registro['PEDIMENTO'] = producto.get('pedimento')
                        firma1 = firmas.get('firma1') if isinstance(firmas, dict) else None
                        if firma1:
                            registro['FIRMA'] = firma1.get('codigo_solicitado') or firma1.get('codigo') or firma1.get('nombre')
                        else:
                            registro['FIRMA'] = None

                        registro['DESCRIPCION'] = producto.get('descripcion')
                        if tabla_prod and isinstance(tabla_prod, list):
                            registro['MARCA'] = tabla_prod[0].get('marca')
                            # Recopilar todos los códigos de tabla_productos
                            codigos = []
                            for p in tabla_prod:
                                try:
                                    c = p.get('codigo')
                                except Exception:
                                    c = None
                                if c is not None:
                                    codigos.append(str(c))
                            # Guardar lista de códigos y también un primer código por compatibilidad
                            registro['CODIGOS'] = codigos
                            registro['CODIGO'] = codigos[0] if codigos else None

                        if key in dictamenes_map:
                            # Merge: evitar sobrescribir, agregar registros y no duplicar
                            existing = dictamenes_map[key]
                            # añadir registro si no está ya
                            if registro not in existing.get('registros', []):
                                existing.setdefault('registros', []).append(registro)
                            # mantener dictamen_json existente, pero si falta, asignar
                            if not existing.get('dictamen_json'):
                                existing['dictamen_json'] = d
                            # mantener cadena y sol_token si faltan
                            existing.setdefault('cadena_identificacion', cadena)
                            existing.setdefault('sol_token', sol_token)
                        else:
                            dictamenes_map[key] = {
                                'solicitud': sol,
                                'folio': fol,
                                'registros': [registro],
                                'dictamen_json': d,
                                'cadena_identificacion': cadena,
                                'sol_token': sol_token,
                            }
            except Exception:
                pass

            # Fusionar información proveniente de tabla_de_relacion (si existe), agregando registros
            try:
                for registro in self.tabla_relacion:
                    solicitud_tr = str(registro.get('SOLICITUD', '')).strip()
                    folio_tr = str(registro.get('FOLIO', '')).strip()
                    key_tr = f"{solicitud_tr}_{folio_tr}"

                    # Si existe un dictamen JSON que corresponda a este registro, lo omitimos
                    # porque ya generaremos la fila desde el JSON (evita duplicados y firmas distintas).
                    try:
                        found = self._find_dictamen(solicitud_tr, folio_tr)
                        if found:
                            # hay un dictamen JSON para este registro -> omitir
                            continue
                    except Exception:
                        pass

                    # Si la clave ya existe en el mapa y no proviene de un JSON, anexar
                    if key_tr in dictamenes_map:
                        if not dictamenes_map[key_tr].get('dictamen_json'):
                            if registro not in dictamenes_map[key_tr]['registros']:
                                dictamenes_map[key_tr]['registros'].append(registro)
                    else:
                        # crear nuevo entry por solicitud_folio (registro sin JSON asociado)
                        dictamenes_map[key_tr] = {
                            'solicitud': solicitud_tr,
                            'folio': folio_tr,
                            'registros': [registro]
                        }
            except Exception:
                pass

            dictamenes = list(dictamenes_map.values())
            # Ordenar dictámenes por folio numérico ascendente (usar 'folio' o extraer de 'sol_token')
            def _folio_key(item: Dict) -> int:
                try:
                    return int(item.get('folio') or 0)
                except Exception:
                    # intentar extraer número al final de sol_token o cadena_identificacion
                    try:
                        s = str(item.get('sol_token') or item.get('cadena_identificacion') or '')
                        m = re.search(r"-(\d+)$", s)
                        if m:
                            return int(m.group(1))
                    except Exception:
                        pass
                return 0

            dictamenes.sort(key=_folio_key)
            print(f"📊 Dictámenes encontrados (antes de deduplicar por folio): {len(dictamenes)}")

            # Consolidar para que cada FOLIO EMA sea único.
            # Preferir dictámenes que tengan 'dictamen_json' cuando haya conflicto.
            folio_map: Dict[str, Dict] = {}
            for d in dictamenes:
                fol_raw = d.get('folio')
                fol_key = self.formatear_folio_ema(fol_raw)
                if fol_key not in folio_map:
                    folio_map[fol_key] = d
                else:
                    # si ya existe, preferimos el que tiene dictamen_json
                    existing = folio_map[fol_key]
                    if not existing.get('dictamen_json') and d.get('dictamen_json'):
                        folio_map[fol_key] = d

            # Crear lista ordenada por folio y escribir solo una fila por folio
            consolidated = list(folio_map.values())
            consolidated.sort(key=_folio_key)

            print(f"📊 Dictámenes únicos por FOLIO EMA a exportar: {len(consolidated)}")

            # Generar filas
            fila_actual = 2
            filas_procesadas = 0

            for dictamen in consolidated:
                fila_datos = self.generar_fila_excel(dictamen)

                # Filtrar por fechas si se especificaron
                if not self.filtrar_por_fechas(fila_datos, fecha_inicio, fecha_fin):
                    continue

                # Escribir datos
                for col, encabezado in enumerate(encabezados, 1):
                    valor = fila_datos.get(encabezado, "N/A")
                    celda = ws.cell(row=fila_actual, column=col, value=valor)
                    celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    celda.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

                fila_actual += 1
                filas_procesadas += 1
            
            # Ajustar ancho de columnas
            for col in range(1, len(encabezados) + 1):
                columna_letra = get_column_letter(col)
                # Anchos específicos según el contenido
                if col == 1:  # NÚMERO DE SOLICITUD
                    ws.column_dimensions[columna_letra].width = 18
                elif col in [2, 6, 16]:  # CLIENTE, PRODUCTO, MODELOS
                    ws.column_dimensions[columna_letra].width = 30
                elif col in [11, 14, 15]:  # FECHAS
                    ws.column_dimensions[columna_letra].width = 15
                else:
                    ws.column_dimensions[columna_letra].width = 20
            
            # Congelar primera fila
            ws.freeze_panes = 'A2'
            
            # Guardar archivo
            wb.save(nombre_archivo)
            
            mensaje = f"✅ Archivo Excel generado exitosamente: {nombre_archivo}\n"
            mensaje += f"   📊 Total de registros: {filas_procesadas}"
            
            if fecha_inicio or fecha_fin:
                mensaje += f"\n   📅 Rango de fechas aplicado: "
                mensaje += f"{fecha_inicio or 'inicio'} a {fecha_fin or 'fin'}"
            
            print(mensaje)
            return True, mensaje
            
        except Exception as e:
            mensaje = f"❌ Error al crear archivo Excel: {e}"
            print(mensaje)
            return False, mensaje

def main():
    """Función principal para ejecutar el script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Generar Control de Folios Anual desde archivos JSON"
    )
    parser.add_argument(
        "--output",
        "-o",
        default="Control_Folios_Anual.xlsx",
        help="Nombre del archivo Excel de salida (default: Control_Folios_Anual.xlsx)"
    )
    parser.add_argument(
        "--fecha-inicio",
        "-fi",
        help="Fecha de inicio para filtrar (formato: YYYY-MM-DD)"
    )
    parser.add_argument(
        "--fecha-fin",
        "-ff",
        help="Fecha de fin para filtrar (formato: YYYY-MM-DD)"
    )
    parser.add_argument(
        "--data-dir",
        "-d",
        default="data",
        help="Directorio donde se encuentran los archivos JSON (default: data)"
    )
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("📊 GENERADOR DE CONTROL DE FOLIOS ANUAL")
    print("=" * 70)
    print()
    
    # Crear instancia del generador
    generador = ControlFoliosAnual(data_dir=args.data_dir)
    
    # Cargar datos
    print("📂 Cargando datos desde archivos JSON...")
    exito, mensaje = generador.cargar_datos()
    
    if not exito:
        print(f"\n❌ Error: {mensaje}")
        return 1
    
    print()
    
    # Generar Excel
    exito, mensaje = generador.crear_excel(
        args.output,
        fecha_inicio=args.fecha_inicio,
        fecha_fin=args.fecha_fin
    )
    
    if not exito:
        print(f"\n❌ Error: {mensaje}")
        return 1
    
    print()
    print("=" * 70)
    print("✅ PROCESO COMPLETADO")
    print("=" * 70)
    
    return 0

def generar_control_folios_anual(
    historial_path,
    tabla_backups_dir,
    output_path,
    year,
    start_date=None,
    end_date=None,
    export_cache=None,
    historial_list: Optional[List[Dict]] = None,
    data_dir: Optional[str] = None
):
    from datetime import datetime
    import os

    def normalizar(fecha):
        if not fecha:
            return None
        try:
            return datetime.strptime(fecha, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            return None

    fecha_inicio = normalizar(start_date)
    fecha_fin = normalizar(end_date)

    # DEBUG: imprimir parámetros recibidos
    try:
        print(f"[DEBUG] generar_control_folios_anual called with:\n  historial_path={historial_path}\n  tabla_backups_dir={tabla_backups_dir}\n  output_path={output_path}\n  year={year}\n  start_date={start_date} -> {fecha_inicio}\n  end_date={end_date} -> {fecha_fin}\n  export_cache={export_cache}\n")
    except Exception:
        pass

    # Resolver data_dir: si se pasó explícitamente lo usamos, si no
    # lo calculamos desde el historial_path como antes.
    if not data_dir:
        base_dir = os.path.dirname(os.path.dirname(historial_path))
        data_dir = os.path.join(base_dir, "data")

    try:
        print(f"[DEBUG] Resolved data_dir: {data_dir}")
        dicts_dir = os.path.join(data_dir, 'Dictamenes')
        if os.path.exists(dicts_dir):
            files = [f for f in os.listdir(dicts_dir) if f.lower().endswith('.json')]
        else:
            files = []
        print(f"[DEBUG] Dictamenes files in {dicts_dir}: {len(files)}")
    except Exception:
        pass

    generador = ControlFoliosAnual(data_dir=data_dir)

    exito, mensaje = generador.cargar_datos()
    if not exito:
        raise Exception(mensaje)

    # Si recibimos una lista de historial (p.e. desde export_cache), usarla
    # en lugar de leer el `historial_visitas.json` del data_dir. Luego
    # regenerar el mapeo folio->cliente.
    if historial_list is not None:
        # historial_list puede venir como {'visitas': [...]} o directamente como lista
        if isinstance(historial_list, dict) and 'visitas' in historial_list:
            generador.historial_visitas = historial_list['visitas']
        elif isinstance(historial_list, list):
            generador.historial_visitas = historial_list
        else:
            # ignorar si no tiene formato esperado
            pass
        # recrear mapeo
        try:
            generador._crear_mapeo_folio_cliente()
        except Exception:
            pass

    exito, mensaje = generador.crear_excel(
        output_path,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin
    )

    if not exito:
        raise Exception(mensaje)

    return True


def generar_reporte_ema(tabla_de_relacion_path, historial_path, output_path, export_cache=None):
    """Genera el reporte EMA a partir de un archivo de tabla_de_relacion (o lista JSON).

    Args:
        tabla_de_relacion_path: Ruta al JSON de tabla_de_relacion o a un JSON temporal con registros.
        historial_path: Ruta al historial (se usa para resolver data_dir y clientes si es necesario).
        output_path: Ruta de salida .xlsx
        export_cache: (opcional) ruta a cache de export
    """
    import os
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Resolver data_dir igual que en generar_control_folios_anual
    base_dir = os.path.dirname(os.path.dirname(historial_path))
    data_dir = os.path.join(base_dir, "data")

    generador = ControlFoliosAnual(data_dir=data_dir)
    exito, mensaje = generador.cargar_datos()
    if not exito:
        raise Exception(mensaje)

    # Cargar tabla_de_relacion
    if not os.path.exists(tabla_de_relacion_path):
        raise Exception(f"No se encontró tabla_de_relacion: {tabla_de_relacion_path}")

    try:
        with open(tabla_de_relacion_path, 'r', encoding='utf-8') as f:
            tabla_obj = json.load(f)
    except Exception as e:
        raise Exception(f"Error leyendo {tabla_de_relacion_path}: {e}")

    # Normalizar a lista de registros
    if isinstance(tabla_obj, dict):
        # Si es dict y contiene una lista en alguna clave esperada
        if 'tabla' in tabla_obj and isinstance(tabla_obj['tabla'], list):
            registros = tabla_obj['tabla']
        elif 'registros' in tabla_obj and isinstance(tabla_obj['registros'], list):
            registros = tabla_obj['registros']
        else:
            # si es un dict que ya representa una lista envuelta
            registros = []
            for v in tabla_obj.values():
                if isinstance(v, list):
                    registros = v
                    break
    elif isinstance(tabla_obj, list):
        registros = tabla_obj
    else:
        raise Exception("Formato de tabla_de_relacion no reconocido")

    # Agrupar por SOLICITUD + FOLIO
    grupos = {}
    for reg in registros:
        solicitud = reg.get('SOLICITUD', '')
        folio = reg.get('FOLIO', '')
        clave = f"{solicitud}_{folio}"
        grupos.setdefault(clave, []).append(reg)

    # Preparar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "EMA"

    encabezados = [
        "Número de solicitud",
        "Fecha de inspección",
        "Número de dictamen",
        "Número de Contrato",
        "Tipo de Documento Oficial Emitido",
        "Fecha de Documento Emitido",
        "Producto verificado",
        "Fecha de Desaduanamiento",
        "Fecha de visita",
        "Observaciones",
        "Inspector(es)",
        "Persona(s) de apoyo",
        "NOM"
    ]

    # Escribir encabezados con estilo
    for col, h in enumerate(encabezados, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    fila = 2
    filas_procesadas = 0

    for clave, regs in grupos.items():
        primer = regs[0]
        solicitud = primer.get('SOLICITUD', '')
        folio = primer.get('FOLIO', '')

        # Formatos y búsquedas
        try:
            folio_num = int(folio) if folio not in (None, '') else 0
        except Exception:
            folio_num = 0

        cliente_info = generador.buscar_cliente_por_solicitud(solicitud, folio_num)

        # Construir campos
        numero_solicitud = generador.extraer_sol_ema(solicitud)
        fecha_inspeccion = primer.get('FECHA DE VERIFICACION', 'N/A')
        numero_dictamen = generador.formatear_folio_ema(folio)
        numero_contrato = cliente_info.get('NÚMERO_DE_CONTRATO', 'N/A') if cliente_info else 'N/A'
        tipo_doc = primer.get('TIPO DE DOCUMENTO', primer.get('TIPO DE DOCUMENTO OFICIAL EMITIDO', 'D'))
        fecha_doc_emitido = primer.get('FECHA DE EMISION DE SOLICITUD', 'N/A')

        # Productos, noms
        productos = set()
        noms = set()
        for r in regs:
            if r.get('DESCRIPCION'):
                productos.add(r.get('DESCRIPCION'))
            if r.get('CLASIF UVA'):
                noms.add(str(r.get('CLASIF UVA')))

        producto_verificado = ", ".join(productos) if productos else 'N/A'
        fecha_desaduanamiento = primer.get('FECHA DE ENTRADA', 'N/A')
        fecha_visita = primer.get('FECHA DE VERIFICACION', 'N/A')
        observaciones = 'N/A'

        # Inspector(es)
        firma = primer.get('FIRMA', '')
        inspector_nombre = generador.buscar_inspector_por_firma(firma)
        inspector_nombre = generador._normalize_name(inspector_nombre)

        personas_apoyo = 'N/A'
        nom_str = ", ".join(noms) if noms else 'N/A'

        fila_vals = [
            numero_solicitud,
            fecha_inspeccion,
            numero_dictamen,
            numero_contrato,
            tipo_doc,
            fecha_doc_emitido,
            producto_verificado,
            fecha_desaduanamiento,
            fecha_visita,
            observaciones,
            inspector_nombre,
            personas_apoyo,
            nom_str
        ]

        for col, val in enumerate(fila_vals, 1):
            cel = ws.cell(row=fila, column=col, value=val)
            cel.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cel.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        fila += 1
        filas_procesadas += 1

    # Ajustar anchos de columna
    for col in range(1, len(encabezados) + 1):
        col_letter = get_column_letter(col)
        if col in (1, 3, 4):
            ws.column_dimensions[col_letter].width = 18
        elif col in (2, 7):
            ws.column_dimensions[col_letter].width = 30
        else:
            ws.column_dimensions[col_letter].width = 20

    ws.freeze_panes = 'A2'

    try:
        wb.save(output_path)
    except Exception as e:
        raise Exception(f"Error guardando Excel EMA: {e}")

    return True

if __name__ == "__main__":
    exit(main())

