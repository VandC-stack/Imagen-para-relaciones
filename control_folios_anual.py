#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para generar el Control de Folios Anual en formato Excel
Autor: Sistema de generación de dictámenes
Fecha: Diciembre 2024

Este script lee datos de múltiples archivos JSON y genera un archivo Excel
con el control de folios anual según los requisitos especificados.
"""

import json
import os
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ControlFoliosAnual:
    """Clase para generar el control de folios anual desde archivos JSON"""
    
    def __init__(self, data_dir: str = "data"):
        """
        Inicializar el generador de control de folios
        
        Args:
            data_dir: Directorio donde se encuentran los archivos JSON
        """
        self.data_dir = data_dir
        self.clientes = []
        self.firmas = []
        self.tabla_relacion = []
        
    def cargar_datos(self) -> Tuple[bool, str]:
        """
        Cargar todos los archivos JSON necesarios
        
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            # Cargar Clientes.json
            clientes_path = os.path.join(self.data_dir, "Clientes.json")
            if os.path.exists(clientes_path):
                with open(clientes_path, 'r', encoding='utf-8') as f:
                    self.clientes = json.load(f)
                print(f"✅ Clientes cargados: {len(self.clientes)} registros")
            else:
                return False, f"No se encontró {clientes_path}"
            
            # Cargar Firmas.json
            firmas_path = os.path.join(self.data_dir, "Firmas.json")
            if os.path.exists(firmas_path):
                with open(firmas_path, 'r', encoding='utf-8') as f:
                    self.firmas = json.load(f)
                print(f"✅ Firmas cargadas: {len(self.firmas)} registros")
            else:
                return False, f"No se encontró {firmas_path}"
            
            # Cargar tabla_de_relacion.json
            tabla_path = os.path.join(self.data_dir, "tabla_de_relacion.json")
            if os.path.exists(tabla_path):
                with open(tabla_path, 'r', encoding='utf-8') as f:
                    self.tabla_relacion = json.load(f)
                print(f"✅ Tabla de relación cargada: {len(self.tabla_relacion)} registros")
            else:
                return False, f"No se encontró {tabla_path}"
            
            return True, "Datos cargados correctamente"
            
        except json.JSONDecodeError as e:
            return False, f"Error al decodificar JSON: {e}"
        except Exception as e:
            return False, f"Error al cargar datos: {e}"
    
    def buscar_cliente_por_solicitud(self, solicitud: str) -> Optional[Dict]:
        """
        Buscar información del cliente basándose en el número de solicitud
        
        Args:
            solicitud: Número de solicitud (ej: "006916/25")
            
        Returns:
            Diccionario con información del cliente o None
        """
        # En un caso real, necesitaríamos un mapping entre solicitud y cliente
        # Por ahora, retornamos el primer cliente si existe
        if self.clientes:
            return self.clientes[0]
        return None
    
    def buscar_inspector_por_firma(self, firma: str) -> str:
        """
        Buscar el nombre completo del inspector por su firma
        
        Args:
            firma: Código de firma (ej: "GRAMIREZ")
            
        Returns:
            Nombre completo del inspector o "N/A"
        """
        for inspector in self.firmas:
            if inspector.get("FIRMA") == firma:
                return inspector.get("NOMBRE DE INSPECTOR", "N/A")
        return "N/A"
    
    def formatear_folio_ema(self, folio) -> str:
        """
        Formatear el folio EMA a 6 dígitos
        
        Args:
            folio: Número de folio
            
        Returns:
            Folio formateado a 6 dígitos
        """
        try:
            folio_str = str(int(folio))
            return folio_str.zfill(6)
        except (ValueError, TypeError):
            return "000000"
    
    def extraer_sol_ema(self, numero_solicitud: str) -> str:
        """
        Extraer los últimos valores del número de solicitud
        
        Args:
            numero_solicitud: Número de solicitud completo (ej: "006916/25")
            
        Returns:
            Últimos valores separados por guión
        """
        if not numero_solicitud:
            return "N/A"
        # Extraer los últimos componentes separados por '/'
        partes = numero_solicitud.split('/')
        if len(partes) >= 2:
            return f"{partes[-2]}-{partes[-1]}"
        return numero_solicitud
    
    def agrupar_por_dictamen(self) -> List[Dict]:
        """
        Agrupar los registros de tabla_relacion por dictamen (SOLICITUD + FOLIO)
        
        Returns:
            Lista de dictámenes agrupados con su información
        """
        dictamenes = {}
        
        for registro in self.tabla_relacion:
            solicitud = registro.get("SOLICITUD", "")
            folio = registro.get("FOLIO", "")
            
            # Crear clave única por dictamen
            clave_dictamen = f"{solicitud}_{folio}"
            
            if clave_dictamen not in dictamenes:
                dictamenes[clave_dictamen] = {
                    "solicitud": solicitud,
                    "folio": folio,
                    "registros": []
                }
            
            dictamenes[clave_dictamen]["registros"].append(registro)
        
        return list(dictamenes.values())
    
    def generar_fila_excel(self, dictamen: Dict) -> Dict:
        """
        Generar una fila del Excel a partir de un dictamen
        
        Args:
            dictamen: Diccionario con información del dictamen
            
        Returns:
            Diccionario con los datos de la fila
        """
        registros = dictamen["registros"]
        primer_registro = registros[0] if registros else {}
        
        solicitud = dictamen["solicitud"]
        folio = dictamen["folio"]
        
        # Buscar cliente (por ahora usamos el primero disponible)
        cliente = self.buscar_cliente_por_solicitud(solicitud)
        
        # Obtener información del inspector
        firma = primer_registro.get("FIRMA", "")
        nombre_inspector = self.buscar_inspector_por_firma(firma)
        
        # Extraer descripciones, marcas, NOMs y modelos de todos los registros
        descripciones = set()
        marcas = set()
        noms = set()
        modelos = []
        
        for reg in registros:
            if reg.get("DESCRIPCION"):
                descripciones.add(reg.get("DESCRIPCION"))
            if reg.get("MARCA"):
                marcas.add(reg.get("MARCA"))
            if reg.get("CLASIF UVA"):
                noms.add(str(reg.get("CLASIF UVA")))
            if reg.get("CODIGO"):
                modelos.append(str(reg.get("CODIGO")))
        
        # Construir fila
        fila = {
            "NÚMERO DE SOLICITUD": solicitud or "N/A",
            "CLIENTE": cliente.get("CLIENTE", "N/A") if cliente else "N/A",
            "NÚMERO DE CONTRATO": cliente.get("NÚMERO_DE_CONTRATO", "N/A") if cliente else "N/A",
            "RFC": cliente.get("RFC", "N/A") if cliente else "N/A",
            "CURP": "N/A",
            "PRODUCTO VERIFICADO": ", ".join(descripciones) if descripciones else "N/A",
            "MARCAS": ", ".join(marcas) if marcas else "N/A",
            "NOM": ", ".join(noms) if noms else "N/A",
            "TIPO DE DOCUMENTO OFICIAL EMITIDO": "D",
            "DOCUMENTO EMITIDO": solicitud or "N/A",
            "FECHA DE DOCUMENTO EMITIDO": primer_registro.get("FECHA DE EMISION DE SOLICITUD", "N/A"),
            "VERIFICADOR": nombre_inspector,
            "PEDIMENTO DE IMPORTACION": primer_registro.get("PEDIMENTO", "N/A"),
            "FECHA DE DESADUANAMIENTO (CUANDO APLIQUE)": primer_registro.get("FECHA DE ENTRADA", "N/A"),
            "FECHA DE VISITA (CUANDO APLIQUE)": primer_registro.get("FECHA DE VERIFICACION", "N/A"),
            "MODELOS": ", ".join(modelos) if modelos else "N/A",
            "SOL EMA": self.extraer_sol_ema(solicitud),
            "FOLIO EMA": self.formatear_folio_ema(folio),
            "INSP EMA": nombre_inspector
        }
        
        return fila
    
    def filtrar_por_fechas(self, fila: Dict, fecha_inicio: Optional[str] = None, 
                          fecha_fin: Optional[str] = None) -> bool:
        """
        Filtrar una fila por rango de fechas
        
        Args:
            fila: Fila de datos
            fecha_inicio: Fecha de inicio en formato YYYY-MM-DD
            fecha_fin: Fecha de fin en formato YYYY-MM-DD
            
        Returns:
            True si la fila está en el rango, False si no
        """
        if not fecha_inicio and not fecha_fin:
            return True
        
        # Usar la fecha de verificación para filtrar
        fecha_str = fila.get("FECHA DE VISITA (CUANDO APLIQUE)", "")
        
        if not fecha_str or fecha_str == "N/A":
            return False
        
        try:
            # Intentar parsear la fecha en diferentes formatos
            for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"]:
                try:
                    fecha = datetime.strptime(fecha_str, fmt)
                    break
                except ValueError:
                    continue
            else:
                # Si no se pudo parsear, incluir el registro
                return True
            
            if fecha_inicio:
                inicio = datetime.strptime(fecha_inicio, "%Y-%m-%d")
                if fecha < inicio:
                    return False
            
            if fecha_fin:
                fin = datetime.strptime(fecha_fin, "%Y-%m-%d")
                if fecha > fin:
                    return False
            
            return True
            
        except Exception:
            # En caso de error, incluir el registro
            return True
    
    def crear_excel(self, nombre_archivo: str, fecha_inicio: Optional[str] = None,
                   fecha_fin: Optional[str] = None) -> Tuple[bool, str]:
        """
        Crear el archivo Excel con el control de folios
        
        Args:
            nombre_archivo: Nombre del archivo Excel a crear
            fecha_inicio: Fecha de inicio para filtrar (YYYY-MM-DD)
            fecha_fin: Fecha de fin para filtrar (YYYY-MM-DD)
            
        Returns:
            Tuple[bool, str]: (éxito, mensaje)
        """
        try:
            print("\n🚀 Generando archivo Excel...")
            
            # Crear libro de trabajo
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Control de Folios"
            
            # Definir encabezados
            encabezados = [
                "NÚMERO DE SOLICITUD",
                "CLIENTE",
                "NÚMERO DE CONTRATO",
                "RFC",
                "CURP",
                "PRODUCTO VERIFICADO",
                "MARCAS",
                "NOM",
                "TIPO DE DOCUMENTO OFICIAL EMITIDO",
                "DOCUMENTO EMITIDO",
                "FECHA DE DOCUMENTO EMITIDO",
                "VERIFICADOR",
                "PEDIMENTO DE IMPORTACION",
                "FECHA DE DESADUANAMIENTO (CUANDO APLIQUE)",
                "FECHA DE VISITA (CUANDO APLIQUE)",
                "MODELOS",
                "SOL EMA",
                "FOLIO EMA",
                "INSP EMA"
            ]
            
            # Escribir encabezados
            for col, encabezado in enumerate(encabezados, 1):
                celda = ws.cell(row=1, column=col, value=encabezado)
                # Estilo para encabezados
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                celda.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Agrupar datos por dictamen
            dictamenes = self.agrupar_por_dictamen()
            print(f"📊 Dictámenes encontrados: {len(dictamenes)}")
            
            # Generar filas
            fila_actual = 2
            filas_procesadas = 0
            
            for dictamen in dictamenes:
                fila_datos = self.generar_fila_excel(dictamen)
                
                # Filtrar por fechas si se especificaron
                if not self.filtrar_por_fechas(fila_datos, fecha_inicio, fecha_fin):
                    continue
                
                # Escribir datos
                for col, encabezado in enumerate(encabezados, 1):
                    valor = fila_datos.get(encabezado, "N/A")
                    celda = ws.cell(row=fila_actual, column=col, value=valor)
                    celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    celda.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                
                fila_actual += 1
                filas_procesadas += 1
            
            # Ajustar ancho de columnas
            for col in range(1, len(encabezados) + 1):
                columna_letra = get_column_letter(col)
                # Anchos específicos según el contenido
                if col == 1:  # NÚMERO DE SOLICITUD
                    ws.column_dimensions[columna_letra].width = 18
                elif col in [2, 6, 16]:  # CLIENTE, PRODUCTO, MODELOS
                    ws.column_dimensions[columna_letra].width = 30
                elif col in [11, 14, 15]:  # FECHAS
                    ws.column_dimensions[columna_letra].width = 15
                else:
                    ws.column_dimensions[columna_letra].width = 20
            
            # Congelar primera fila
            ws.freeze_panes = 'A2'
            
            # Guardar archivo
            wb.save(nombre_archivo)
            
            mensaje = f"✅ Archivo Excel generado exitosamente: {nombre_archivo}\n"
            mensaje += f"   📊 Total de registros: {filas_procesadas}"
            
            if fecha_inicio or fecha_fin:
                mensaje += f"\n   📅 Rango de fechas aplicado: "
                mensaje += f"{fecha_inicio or 'inicio'} a {fecha_fin or 'fin'}"
            
            print(mensaje)
            return True, mensaje
            
        except Exception as e:
            mensaje = f"❌ Error al crear archivo Excel: {e}"
            print(mensaje)
            return False, mensaje


def main():
    """Función principal para ejecutar el script"""
    import argparse
    
    parser = argparse.ArgumentParser(
        description="Generar Control de Folios Anual desde archivos JSON"
    )
    parser.add_argument(
        "--output",
        "-o",
        default="Control_Folios_Anual.xlsx",
        help="Nombre del archivo Excel de salida (default: Control_Folios_Anual.xlsx)"
    )
    parser.add_argument(
        "--fecha-inicio",
        "-fi",
        help="Fecha de inicio para filtrar (formato: YYYY-MM-DD)"
    )
    parser.add_argument(
        "--fecha-fin",
        "-ff",
        help="Fecha de fin para filtrar (formato: YYYY-MM-DD)"
    )
    parser.add_argument(
        "--data-dir",
        "-d",
        default="data",
        help="Directorio donde se encuentran los archivos JSON (default: data)"
    )
    
    args = parser.parse_args()
    
    print("=" * 70)
    print("📊 GENERADOR DE CONTROL DE FOLIOS ANUAL")
    print("=" * 70)
    print()
    
    # Crear instancia del generador
    generador = ControlFoliosAnual(data_dir=args.data_dir)
    
    # Cargar datos
    print("📂 Cargando datos desde archivos JSON...")
    exito, mensaje = generador.cargar_datos()
    
    if not exito:
        print(f"\n❌ Error: {mensaje}")
        return 1
    
    print()
    
    # Generar Excel
    exito, mensaje = generador.crear_excel(
        args.output,
        fecha_inicio=args.fecha_inicio,
        fecha_fin=args.fecha_fin
    )
    
    if not exito:
        print(f"\n❌ Error: {mensaje}")
        return 1
    
    print()
    print("=" * 70)
    print("✅ PROCESO COMPLETADO")
    print("=" * 70)
    
    return 0


if __name__ == "__main__":
    exit(main())
